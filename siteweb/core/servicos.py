# siteweb/core/servicos.py
from playwright.sync_api import sync_playwright
from siteweb.core.raspagem import raspar_dados
from utils.cache import salvar_cache, ler_cache
from siteweb.core.utils import limpar_preco
from django.db import models
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings
from datetime import datetime
import random
import time
import json
from django.http import JsonResponse
import pandas as pd
import numpy as np
from siteweb.models import Loja, Produto

def executar_raspagem(urlLoja, cepEntrega):
    print("// -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- //")
    print("INICIO RASPAGEM")

    # 1. Verifica se a loja já existe no DB
    nloja = Loja.objects.filter(url=urlLoja).first()
    if nloja:
        nomeLoja = nloja.nome
        print(f"Loja encontrada no DB: {nomeLoja}")
    else:
        # Se não existe, tenta descobrir via Playwright
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=False)
            USER_AGENTS = [
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/117 Safari/537.36",
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/537.36",
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/116 Safari/537.36"
            ]
            context = browser.new_context(
                user_agent=random.choice(USER_AGENTS),
                viewport={"width": 1280, "height": 800}
            )
            page = context.new_page()
            print("[INFO] Acessando página...")
            page.goto(urlLoja, timeout=50000)
            time.sleep(2)
            page.mouse.wheel(0, 1000)
            time.sleep(1)

            if "algo deu errado" in page.content():
                print("[ERRO] Página de erro detectada. Abortando raspagem.")
                browser.close()
                return [], None, None

            selected_text = page.evaluate("""
                () => {
                    const selectElement = document.querySelector('select');
                    return selectElement ? selectElement.options[selectElement.selectedIndex].text : "Loja Desconhecida";
                }
            """)
            browser.close()
            nomeLoja = selected_text.strip()
            print(f"Nome da loja obtido: {nomeLoja}")

    # 2. Cache
    chave_cache = f"loja:{nomeLoja}"
    produtos_json = ler_cache(chave_cache)
    produtos = json.loads(produtos_json) if produtos_json else []

    if not produtos:
        print("Produtos não encontrados no cache")
        loja = Loja.objects.filter(url=urlLoja).first()

        if loja:
            print("Loja encontrada no DB.")
            produtos_db = Produto.objects.filter(loja_id=loja.id).all()

            if not produtos_db.exists():
                print("[INFO] Produtos não encontrados no DB, iniciando raspagem.")
                with sync_playwright() as playwright:
                    produtos = raspar_dados(playwright, urlLoja, cepEntrega)

                # Normaliza lista de produtos (dicts)
                produto_list = [
                    {
                        "codigoProduto": p["codigoProduto"] if isinstance(p, dict) else p.codigoProduto,
                        "nome": p["nome"] if isinstance(p, dict) else p.nome,
                        "preco": (
                            float(
                                (p["preco"] if isinstance(p, dict) else p.preco)
                                .replace("R$", "")
                                .replace("\xa0", "")
                                .replace(",", ".")
                            )
                            if (p["preco"] if isinstance(p, dict) else p.preco)
                            else None
                        ),
                        "imagem": p["imagem"] if isinstance(p, dict) else (p.imagem if p.imagem else None),
                        "loja": p.get("loja") if isinstance(p, dict) else p.loja_id,
                        "url": p["url"] if isinstance(p, dict) else (p.url if p.url else "teste"),
                        "frete": p["frete"] if isinstance(p, dict) else (p.frete if p.frete else "frete"),
                        "prazo": p["prazo"] if isinstance(p, dict) else (p.prazo if p.prazo else "prazo"),
                    }
                    for p in produtos
                ]
                salvar_cache(chave_cache, json.dumps(produto_list))
                produtos = produto_list
            else:
                print("Produtos encontrados no DB: ", len(produtos_db))
                produto_list = [
                    {
                        "codigoProduto": p.codigoProduto,
                        "nome": p.nome,
                        "preco": float(p.preco),
                        "imagem": p.imagem if p.imagem else None,
                        "loja": p.loja_id,
                        "url": p.url if p.url else "teste",
                        "frete": p.frete ,
                        "prazo": p.prazo ,  
                    }
                    for p in produtos_db
                ]
                salvar_cache(chave_cache, json.dumps(produto_list))
                produtos = produto_list
        else:
            print("Loja não encontrada, raspando direto...")
            with sync_playwright() as playwright:
                produtos = raspar_dados(playwright, urlLoja, cepEntrega)
    else:
        print("Produtos encontrados no cache.")

    print("// -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- -- //")
    return produtos, nomeLoja, chave_cache
           

def salvar_excel_colunas(resultados_sim, destino):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    # estilos
    titulo_font = Font(bold=True, size=12)
    cabecalho_font = Font(bold=True, color="000000")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    formato_moeda = "R$ #,##0.00"
    destaque_principal = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    fundo_cabecalho = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fundo_linha_par = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # agora adicionamos a coluna Ranking como primeira
    colunas = ["ranking", "nome", "preco", "frete", "prazo", "quantidadeCompras", "tipo", "loja", "link"]
    labels = ["Ranking", "Nome", "Preço", "Frete", "Prazo", "Qtd. Compras", "Tipo", "Loja", "Link"]

    linha_atual = 1

    for nomePrincipal, dados in resultados_sim.items():
        # 🔎 verifica se é DataFrame ou dict
        if isinstance(dados, pd.DataFrame):
            principal_records = dados[dados["tipo"] == "principal"].to_dict("records")
            concorrente_records = dados[dados["tipo"] == "concorrente"].to_dict("records")
            principal = principal_records[0] if principal_records else {}
            concorrentes = concorrente_records
        else:
            principal = dados.get("principal", {})
            concorrentes = dados.get("concorrentes", [])

        lista = [principal] + concorrentes
        lista = [item for item in lista if item.get("preco") is not None]
        lista.sort(key=lambda x: x["preco"])

        # título
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(colunas))
        cell = ws.cell(row=linha_atual, column=1, value=principal.get("nome", nomePrincipal))
        cell.font = titulo_font
        cell.alignment = alinhamento_centro
        cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        linha_atual += 1

        # cabeçalho
        for j, label in enumerate(labels, start=1):
            hcell = ws.cell(row=linha_atual, column=j, value=label)
            hcell.font = cabecalho_font
            hcell.alignment = alinhamento_centro
            hcell.fill = fundo_cabecalho
            hcell.border = borda_fina
        linha_atual += 1

        # dados com ranking
        ranking = 1
        for item in lista:
            for j, key in enumerate(colunas, start=1):
                if key == "ranking":
                    val = ranking
                else:
                    val = item.get(key, "")

                cell = ws.cell(row=linha_atual, column=j, value=val)

                if key == "preco" and val is not None:
                    cell.number_format = formato_moeda
                    cell.alignment = alinhamento_centro
                elif key in ["tipo", "quantidadeCompras", "ranking"]:
                    cell.alignment = alinhamento_centro
                    if key == "quantidadeCompras" and val not in ("", None):
                        cell.number_format = "0"
                elif key == "link" and val:
                    cell.hyperlink = val
                    cell.style = "Hyperlink"
                    cell.alignment = alinhamento_esquerda
                else:
                    cell.alignment = alinhamento_esquerda

                cell.border = borda_fina

                # zebra striping
                if linha_atual % 2 == 0 and item.get("tipo") != "principal":
                    cell.fill = fundo_linha_par

            if item.get("tipo") == "principal":
                for j in range(1, len(colunas) + 1):
                    ws.cell(row=linha_atual, column=j).fill = destaque_principal

            linha_atual += 1
            ranking += 1  # incrementa ranking para cada produto

        linha_atual += 1

    # largura dinâmica
    for col_idx, col_name in enumerate(colunas, start=1):
        max_length = max(
            (len(str(ws.cell(row=row, column=col_idx).value)) for row in range(1, ws.max_row+1)),
            default=0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 5

    wb.save(destino)
    print(f"✅ Excel salvo em: {destino}")



def rankeador(resultados_sim):
    planilhas = {}
    textos = []

    if not resultados_sim:
        print("Lista de resultados_sim está vazia.")
        return None 

    for nomePrincipal, dados in resultados_sim.items():
        lista = []

        # pega o principal
        principal = dados.get("principal", {})
        if principal:
            lista.append({
                #"codigoProduto":principal.get("codigoProduto"),
                "principal": nomePrincipal,
                "nome": principal.get("nome"),
                "preco": principal.get("preco"),
                "tipo": "principal",
                "loja": principal.get("loja"),
                "link": principal.get("url", "#"),
                "frete": principal.get("frete"),
                "prazo": principal.get("prazo")
            })

        # pega os concorrentes
        for c in dados.get("concorrentes", []):
            frete_c = c.get("frete") or 0
            lista.append({
                "principal": nomePrincipal,
                "nome": c.get("nome"),
                "preco": c.get("preco") + limpar_preco(str(frete_c)),
                "tipo": "concorrente",
                "loja": c.get("loja"),
                "link": c.get("url", "#"),
                "frete": c.get("frete") if principal.get("frete")else print("O rankeador não recebeu Frete"),
                "prazo": c.get("prazo"),
                "quantidadeCompras": c.get("quantidadeCompras")
            })
            print(f"➡️ Concorrente rankeador {c.get('nome')} | Valor: {c.get('preco')} | Qtd Compras: {c.get('quantidadeCompras')}")

        # remove itens sem preço
        lista = [item for item in lista if item["preco"] is not None]
        # ordena por preço
        lista.sort(key=lambda x: x["preco"])

        # encontra índice do principal
        indicePrinci = [i for i, item in enumerate(lista) if item["tipo"] == "principal"]
        texto = "j"
        if indicePrinci:
            idx = indicePrinci[0]
            precoAntigo = lista[idx]["preco"]

            if idx >= 3 and len(lista) > 2:
                valorTerceiro = lista[2]["preco"]
                lista.insert(2, lista.pop(idx)) 
                novo_preco = max(0, valorTerceiro - 1)
                lista[2]["preco"] = novo_preco
                texto = (
                    f"O principal '{nomePrincipal}' foi movido para a 3ª posição; "
                    f"preço ajustado de {precoAntigo} para {novo_preco} para melhorar competitividade."
                )
            else:
                texto = "O produto principal já está entre os três primeiros, sem necessidade de ajuste de preço."
        else:
            texto = "Produto principal não encontrado na lista de concorrentes."

        textos.append(texto)

        # salva ranking desse produto em DataFrame
        planilhas[nomePrincipal] = pd.DataFrame(lista)
    salvar_excel_colunas(planilhas, destino=f"{settings.MEDIA_ROOT}/classificacaoPrecoExcel/Rankeamento-{datetime.now().strftime('%d-%m-%Y_%H%M')}.xlsx")
    planilhas_dict = {nome: df.to_dict(orient="records") for nome, df in planilhas.items()}


    return planilhas_dict, textos